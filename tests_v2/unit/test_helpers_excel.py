from pathlib import Path

import openpyxl
import pytest

from eval.helpers import check_cell_value, check_formatting_conventions


@pytest.mark.unit
def test_check_cell_value_numeric_exact(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"].value = 10
    file_path = tmp_path / "test.xlsx"
    wb.save(file_path)

    passed, actual, details = check_cell_value(file_path, "A1", expected=10)
    assert passed is True
    assert actual == 10
    assert "expected 10" in details


@pytest.mark.unit
def test_check_cell_value_string_case_insensitive(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B2"].value = "Ok"
    file_path = tmp_path / "test.xlsx"
    wb.save(file_path)

    passed, actual, details = check_cell_value(file_path, "B2", expected="ok")
    assert passed is True
    assert actual == "Ok"
    assert "expected 'ok'" in details


@pytest.mark.unit
def test_check_cell_value_missing_sheet(tmp_path):
    wb = openpyxl.Workbook()
    wb.save(tmp_path / "test.xlsx")
    passed, actual, details = check_cell_value(
        tmp_path / "test.xlsx", "A1", expected=1, sheet="Missing"
    )
    assert passed is False
    assert actual is None
    assert "Sheet 'Missing' not found" in details


@pytest.mark.unit
def test_check_formatting_conventions_detects_violation(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    cell = ws["C3"]
    cell.value = 100
    file_path = tmp_path / "format.xlsx"
    wb.save(file_path)

    passed, violations = check_formatting_conventions(file_path, cells=["C3"])
    assert passed is False
    assert any("hardcoded number should be blue" in v for v in violations)
