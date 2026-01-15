"""
Shared utilities for the IB-bench evaluation pipeline.
"""

import hashlib
import json
import re
import time
from dataclasses import dataclass
from datetime import datetime
from functools import wraps
from pathlib import Path
from typing import Any, Literal, TypedDict, cast

import yaml
from dotenv import load_dotenv

# -----------------------------------------------------------------------------
# Type Definitions
# -----------------------------------------------------------------------------

Provider = Literal["anthropic", "openai", "gemini"]
CriterionType = Literal["programmatic", "llm_judge", "human_judge"]
MatchType = Literal["substring_one_of", "regex_pattern"]


class RubricCriterion(TypedDict, total=False):
    """Single criterion in a rubric. All fields optional due to variation."""

    description: str
    type: CriterionType
    match_type: MatchType
    points: int
    accepted_values: list[str]
    valid_patterns: list[str]
    required_elements: list[str]
    forbidden_elements: list[str]
    gates_llm: bool
    search_full_response: bool
    core_concepts: list[str]
    scoring_guide: str


class Rubric(TypedDict, total=False):
    """Rubric structure loaded from rubric.json."""

    task_id: str
    version: str
    total_points: int
    criteria: dict[str, RubricCriterion]


class TaskMeta(TypedDict, total=False):
    """Task metadata from meta.yaml -> task section."""

    id: str
    title: str
    type: str
    category: str | list[str]
    input_type: str
    description: str


load_dotenv(Path(__file__).parent.parent / ".env")


class JudgeParseError(Exception):
    def __init__(self, message: str, raw_response: str = ""):
        super().__init__(message)
        self.raw_response = raw_response


def retry_on_rate_limit(max_retries: int = 3, initial_wait: int = 60):
    """Decorator to retry on rate limit errors with exponential backoff.

    Args:
        max_retries: Maximum number of retry attempts
        initial_wait: Initial wait time in seconds (doubles each retry)
    """

    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            last_exception = None
            wait_time = initial_wait

            for attempt in range(max_retries + 1):
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    error_str = str(e)
                    # Check if it's a rate limit error (429)
                    if "429" in error_str or "rate_limit" in error_str.lower():
                        last_exception = e
                        if attempt < max_retries:
                            print(
                                f"  Rate limited. Waiting {wait_time}s before retry ({attempt + 1}/{max_retries})..."
                            )
                            time.sleep(wait_time)
                            wait_time *= 2  # Exponential backoff
                        else:
                            print(
                                f"  Rate limited. Max retries ({max_retries}) exceeded."
                            )
                            raise
                    else:
                        # Not a rate limit error, raise immediately
                        raise

            if last_exception is not None:
                raise last_exception
            raise RuntimeError("Retry loop completed without success or exception")

        return wrapper

    return decorator


@dataclass
class Task:
    """Represents a single evaluation task."""

    id: str
    task_dir: Path
    task_type: str
    category: str | list[str]
    description: str
    prompt: str
    rubric: Rubric
    input_files: list[Path]


def load_task(task_dir: Path, include_rubric: bool = True) -> Task:
    """Load a single task from a directory."""
    # Parse meta.yaml
    meta_path = task_dir / "meta.yaml"
    with open(meta_path) as f:
        meta = yaml.safe_load(f)

    # Skip if meta.yaml is not a proper dict (e.g., just a comment blurb)
    if not isinstance(meta, dict):
        raise ValueError(
            f"meta.yaml is not a valid task definition (got {type(meta).__name__})"
        )

    task_meta = meta.get("task", {})

    # Read prompt
    prompt_path = task_dir / "prompt.md"
    prompt = prompt_path.read_text()

    # Read rubric (only if needed for scoring)
    rubric: Rubric = cast(Rubric, {})
    if include_rubric:
        rubric_path = task_dir / "rubric.json"
        if rubric_path.exists():
            with open(rubric_path) as f:
                rubric = cast(Rubric, json.load(f))

    # Find input files (glob pattern: input*.*)
    input_files = list(task_dir.glob("input*.*"))

    return Task(
        id=task_meta.get("id"),
        task_dir=task_dir,
        task_type=task_meta.get("type"),
        category=task_meta.get("category"),
        description=task_meta.get("description", ""),
        prompt=prompt,
        rubric=rubric,
        input_files=input_files,
    )


def load_tasks(
    tasks_dir: Path | None = None,
    task_ids: list[str] | None = None,
    filter_pattern: str | None = None,
    include_rubric: bool = True,
) -> list[Task]:
    """Discover and load tasks from the tasks directory.

    Args:
        tasks_dir: Path to tasks directory. Defaults to eval/tasks/
        task_ids: Optional list of specific task IDs to load (e.g., ["e-001"])
        filter_pattern: Optional prefix filter (e.g., "e-" for easy tasks)
        include_rubric: Whether to load rubric.json (only needed for scoring)

    Returns:
        List of Task objects.
    """
    if tasks_dir is None:
        tasks_dir = Path(__file__).parent / "tasks"

    tasks = []

    for task_path in sorted(tasks_dir.iterdir()):
        if not task_path.is_dir():
            continue

        task_id = task_path.name

        # Filter by task_ids if provided
        if task_ids and task_id not in task_ids:
            continue

        # Filter by pattern if provided
        if filter_pattern and not task_id.startswith(filter_pattern):
            continue

        try:
            task = load_task(task_path, include_rubric=include_rubric)
            tasks.append(task)
        except (FileNotFoundError, ValueError) as e:
            print(f"Warning: Skipping {task_path.name}: {e}")

    return tasks


def get_rubric_hash(rubric: Rubric) -> str:
    """Generate 8-char hash of rubric content for versioning."""
    content = json.dumps(rubric, sort_keys=True)
    return hashlib.sha256(content.encode()).hexdigest()[:8]


def extract_json(text: str) -> dict[str, Any] | None:
    """Extract JSON object from response text."""
    try:
        return json.loads(text.strip())
    except json.JSONDecodeError:
        pass

    code_block = re.search(r"```(?:json)?\s*(.*?)\s*```", text, re.DOTALL)
    if code_block:
        content = code_block.group(1)
        try:
            return json.loads(content)
        except json.JSONDecodeError:
            missing = content.count("{") - content.count("}")
            if missing > 0:
                try:
                    return json.loads(content + "}" * missing)
                except json.JSONDecodeError:
                    pass

    start = text.find("{")
    if start == -1:
        return None

    depth = 0
    for i, char in enumerate(text[start:], start):
        if char == "{":
            depth += 1
        elif char == "}":
            depth -= 1
            if depth == 0:
                try:
                    return json.loads(text[start : i + 1])
                except json.JSONDecodeError:
                    break

    if depth > 0:
        try:
            return json.loads(text[start:] + "}" * depth)
        except json.JSONDecodeError:
            pass

    return None


def extract_task_section(prompt: str) -> str:
    """Extract the ## Task section from a prompt.md file."""
    match = re.search(r"## Task\s*\n(.*?)(?=\n## |\Z)", prompt, re.DOTALL)
    if not match:
        raise ValueError("Prompt missing required '## Task' section")
    return match.group(1).strip()


def _truncate_text(text: str, max_len: int = 300) -> str:
    if len(text) <= max_len:
        return text
    return text[: max_len - 3] + "..."


def extract_error_details(error: Exception) -> dict[str, Any]:
    """Extract useful, JSON-serializable error details from API exceptions."""
    details: dict[str, Any] = {
        "error_type": error.__class__.__name__,
        "message": str(error) or repr(error),
    }

    for attr in ("status_code", "status", "code", "type", "request_id", "param"):
        value = getattr(error, attr, None)
        if value not in (None, ""):
            details[attr] = value

    retryable = getattr(error, "retryable", None)
    if retryable is not None:
        details["retryable"] = retryable

    body = getattr(error, "body", None)
    if body not in (None, ""):
        details["response_body"] = body

    response = getattr(error, "response", None)
    if response is not None:
        status_code = getattr(response, "status_code", None)
        if status_code and "status_code" not in details:
            details["status_code"] = status_code

        headers = getattr(response, "headers", None)
        if headers and "request_id" not in details:
            request_id = headers.get("x-request-id") or headers.get("request-id")
            if request_id:
                details["request_id"] = request_id

        response_text = getattr(response, "text", None)
        if response_text and "response_body" not in details:
            details["response_body"] = response_text

    if "response_body" in details:
        details["response_body"] = _truncate_text(str(details["response_body"]), 2000)

    return details


def format_error_summary(details: dict[str, Any], verbose: bool = False) -> str:
    """Build a short, user-facing error summary."""
    error_type = details.get("error_type", "Error")
    message = details.get("message", "")
    if not verbose:
        message = _truncate_text(str(message), 240)

    parts = [f"{error_type}: {message}" if message else str(error_type)]

    status_code = details.get("status_code") or details.get("status")
    if status_code:
        parts.append(f"status={status_code}")

    error_code = details.get("code") or details.get("error_code")
    if error_code:
        parts.append(f"code={error_code}")

    request_id = details.get("request_id")
    if request_id:
        parts.append(f"request_id={request_id}")

    if verbose and details.get("response_body"):
        parts.append(f"response_body={details['response_body']}")

    return " | ".join(parts)


def suggest_next_steps(details: dict[str, Any]) -> list[str]:
    """Suggest next steps based on error details."""
    message = str(details.get("message", "")).lower()
    status_code = str(details.get("status_code") or details.get("status") or "")

    steps = []

    if "rate limit" in message or status_code == "429":
        steps.append("Reduce --parallel and retry later")
        steps.append("Check provider quota limits")
    elif (
        "api key" in message
        or "authentication" in message
        or status_code
        in {
            "401",
            "403",
        }
    ):
        steps.append("Verify API key environment variables (.env or shell)")
        steps.append("Check account permissions for the model")
    elif "not found" in message and "model" in message:
        steps.append("Verify the model name in the config")
    elif "timeout" in message or "timed out" in message:
        steps.append("Retry the request")
        steps.append("Reduce input size or lower --parallel")
    elif "invalid" in message or status_code == "400":
        steps.append("Inspect task prompt and input files for invalid content")

    if not steps:
        steps.append("Re-run with --verbose to see full error details")

    return steps


def build_error_report(
    error: Exception, verbose: bool
) -> tuple[dict[str, Any], str, list[str]]:
    """Build error details, summary, and next steps for reporting."""
    details = extract_error_details(error)
    summary = format_error_summary(details, verbose=False)
    next_steps = suggest_next_steps(details)
    if verbose:
        next_steps = [step for step in next_steps if "--verbose" not in step]
    return details, summary, next_steps


def get_runner(provider: Provider, model: str):
    """Factory function to get the appropriate runner.

    Uses lazy import to avoid circular dependencies.
    """
    from runners import AnthropicRunner, GeminiRunner, OpenAIRunner

    runners = {
        "anthropic": AnthropicRunner,
        "openai": OpenAIRunner,
        "gemini": GeminiRunner,
    }

    runner_class = runners.get(provider)
    if runner_class is None:
        raise ValueError(f"Unknown provider: {provider}")
    return runner_class(model=model)


def create_run_directory(model: str, base_dir: Path | None = None) -> Path:
    """Create a timestamped run directory under responses/{model}/{run_id}."""
    if base_dir is None:
        base_dir = Path(__file__).parent / "responses"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # Sanitize model name for filesystem
    model_safe = model.replace("/", "-").replace(":", "-")
    run_dir = base_dir / model_safe / timestamp

    run_dir.mkdir(parents=True, exist_ok=True)

    return run_dir


def check_workbook_errors(xlsx_path: Path) -> tuple[bool, list[str]]:
    """
    Check if workbook contains any #REF! or other formula errors.

    :param xlsx_path: Path to Excel file
    :returns: (has_errors, list of error descriptions)
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except Exception as e:
        return True, [f"Failed to open Excel file: {e}"]

    try:
        errors = []
        error_values = {
            "#REF!",
            "#VALUE!",
            "#NAME?",
            "#DIV/0!",
            "#NULL!",
            "#N/A",
            "#NUM!",
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value in error_values:
                        errors.append(f"{sheet_name}!{cell.coordinate}: {cell.value}")

        return len(errors) > 0, errors
    finally:
        wb.close()


def check_cell_value(
    xlsx_path: Path,
    cell: str,
    expected: float | str,
    sheet: str | None = None,
    tolerance: float = 0,
) -> tuple[bool, Any, str]:
    """
    Check if cell equals expected value within tolerance.

    :param xlsx_path: Path to Excel file
    :param cell: Cell reference (e.g., "A3", "B8")
    :param expected: Expected value (numeric with tolerance, or string for exact match)
    :param sheet: Sheet name (defaults to active sheet)
    :param tolerance: Allowed difference for numeric (default 0 = exact match)
    :returns: (passed, actual_value, details)

    Does not handle: Named ranges, formulas (reads computed values only).
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        return False, None, f"Failed to open Excel file: {e}"

    try:
        try:
            ws = wb[sheet] if sheet else wb.active
        except KeyError:
            return False, None, f"Sheet '{sheet}' not found"

        if ws is None:
            return False, None, "No active sheet found"

        actual = ws[cell].value

        if actual is None:
            return False, None, f"Cell {cell} is empty"

        # String comparison (case-insensitive)
        if isinstance(expected, str):
            actual_str = str(actual).strip()
            expected_str = expected.strip()
            passed = actual_str.upper() == expected_str.upper()
            if passed:
                details = f"Cell {cell} = '{actual_str}' (expected '{expected_str}')"
            else:
                details = f"Cell {cell} = '{actual_str}', expected '{expected_str}'"
            return passed, actual_str, details

        # Numeric comparison
        try:
            actual_num = float(actual)
        except (TypeError, ValueError):
            return False, actual, f"Cell {cell} is not numeric: {actual}"

        diff = abs(actual_num - expected)
        passed = diff <= tolerance

        if passed:
            details = f"Cell {cell} = {actual_num} (expected {expected})"
        else:
            details = f"Cell {cell} = {actual_num}, expected {expected} (diff: {diff})"

        return passed, actual_num, details
    finally:
        wb.close()


BLUE_RGB = "0000FF"
GREEN_RGB = "00FF00"
RED_RGB = "FF0000"
BLUE_THEME_INDICES = {4, 5}
GREEN_THEME_INDICES = {6, 9}
RED_THEME_INDICES = {1, 2}


def _is_font_color(font, rgb_suffix: str, theme_indices: set[int]) -> bool:
    """Check if font matches a color by RGB suffix or theme index."""
    if font.color is None:
        return False
    if font.color.rgb and str(font.color.rgb).upper().endswith(rgb_suffix):
        return True
    return font.color.theme in theme_indices


def _is_blue(font) -> bool:
    return _is_font_color(font, BLUE_RGB, BLUE_THEME_INDICES)


def _is_green(font) -> bool:
    return _is_font_color(font, GREEN_RGB, GREEN_THEME_INDICES)


def _is_red(font) -> bool:
    return _is_font_color(font, RED_RGB, RED_THEME_INDICES)


def _has_external_workbook_ref(formula: str) -> bool:
    """Detect [workbook.xlsx]Sheet!Cell syntax."""
    if not formula or not isinstance(formula, str):
        return False
    return "[" in formula and "]" in formula


def _has_cross_sheet_ref(formula: str) -> bool:
    """Detect Sheet!Cell syntax (same workbook, different sheet)."""
    if not formula or not isinstance(formula, str):
        return False
    return "!" in formula and not _has_external_workbook_ref(formula)


def check_formatting_conventions(
    xlsx_path: Path,
    cells: list[str] | None = None,
    sheet: str | None = None,
) -> tuple[bool, list[str]]:
    """
    Check IB Excel formatting conventions:
    - Blue: hardcoded numbers
    - Green: cross-sheet refs (same workbook)
    - Red: external workbook refs

    Known limitations:
    - Theme/tint colors are not resolved; theme indices are heuristic.
    - Conditional formatting colors are not evaluated.
    - Cross-sheet detection relies on explicit "Sheet!Cell" syntax and may miss
      INDIRECT/named ranges.
    - External workbook detection only checks for "[workbook]" syntax.
    """
    import openpyxl

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except Exception as e:
        return False, [f"Failed to open Excel file: {e}"]

    try:
        try:
            ws = wb[sheet] if sheet else wb.active
        except KeyError:
            return False, [f"Sheet '{sheet}' not found"]

        if ws is None:
            return False, ["No active sheet found"]

        violations: list[str] = []

        def check_cell(cell):
            val = cell.value
            font = cell.font

            if val is None:
                return

            is_formula = isinstance(val, str) and val.startswith("=")

            if is_formula:
                if _has_external_workbook_ref(val) and not _is_red(font):
                    violations.append(
                        f"{cell.coordinate}: external workbook ref should be red"
                    )
                elif _has_cross_sheet_ref(val) and not _is_green(font):
                    violations.append(
                        f"{cell.coordinate}: cross-sheet ref should be green"
                    )
            else:
                if isinstance(val, (int, float)) and not _is_blue(font):
                    violations.append(
                        f"{cell.coordinate}: hardcoded number should be blue"
                    )

        if cells:
            for cell_ref in cells:
                if ":" in cell_ref:
                    for row in ws[cell_ref]:
                        for cell in row if hasattr(row, "__iter__") else [row]:
                            check_cell(cell)
                else:
                    check_cell(ws[cell_ref])
        else:
            for row in ws.iter_rows():
                for cell in row:
                    check_cell(cell)

        return len(violations) == 0, violations
    finally:
        wb.close()
