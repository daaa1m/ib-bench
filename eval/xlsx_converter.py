"""
Convert xlsx files to structured JSON preserving formulas and formatting for LLM input.

Usage:
    uv run python eval/xlsx_converter.py path/to/file.xlsx
    uv run python eval/xlsx_converter.py path/to/file.xlsx --output output.json
"""

import json
import sys
from datetime import datetime, date, time
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border


def _get_color_value(color) -> str | None:
    """Extract hex color value from openpyxl color object."""
    if color is None:
        return None
    if color.type == "rgb" and color.rgb:
        # RGB format: AARRGGBB, strip alpha
        rgb = str(color.rgb)
        if len(rgb) == 8:
            return f"#{rgb[2:]}"  # Strip alpha channel
        return f"#{rgb}"
    if color.type == "indexed":
        # Common indexed colors
        indexed_colors = {
            0: "#000000",  # Black
            1: "#FFFFFF",  # White
            2: "#FF0000",  # Red
            3: "#00FF00",  # Green
            4: "#0000FF",  # Blue
            5: "#FFFF00",  # Yellow
            6: "#FF00FF",  # Magenta
            7: "#00FFFF",  # Cyan
        }
        return indexed_colors.get(color.indexed)
    if color.type == "theme":
        return f"theme:{color.theme}"
    return None


def _get_formatting(cell) -> dict | None:
    """Extract formatting information from a cell."""
    fmt = {}

    # Font
    font = cell.font
    if font:
        if font.bold:
            fmt["bold"] = True
        if font.italic:
            fmt["italic"] = True
        if font.underline:
            fmt["underline"] = True
        font_color = _get_color_value(font.color)
        if font_color and font_color != "#000000":
            fmt["font_color"] = font_color

    # Fill/background color
    fill = cell.fill
    if fill and isinstance(fill, PatternFill):
        if fill.patternType and fill.patternType != "none":
            bg_color = _get_color_value(fill.fgColor)
            if bg_color:
                fmt["background"] = bg_color

    # Number format (currency, percentage, date, etc.)
    if cell.number_format and cell.number_format != "General":
        fmt["number_format"] = cell.number_format

    # Alignment
    if cell.alignment:
        if cell.alignment.horizontal and cell.alignment.horizontal != "general":
            fmt["align"] = cell.alignment.horizontal

    return fmt if fmt else None


def xlsx_to_json(filepath: str, include_empty: bool = False, include_formatting: bool = True) -> dict:
    """
    Convert xlsx to structured JSON preserving formulas and formatting.

    Returns:
        {
            "filename": "model.xlsx",
            "sheets": {
                "Sheet1": {
                    "dimensions": "A1:P200",
                    "cells": {
                        "A1": {"value": "Revenue", "type": "string", "format": {"bold": true}},
                        "B1": {"value": 1000, "type": "number", "format": {"font_color": "#0000FF", "number_format": "#,##0"}},
                        "C1": {"formula": "=SUM(A1:B1)", "value": 1500, "type": "formula"}
                    }
                }
            },
            "named_ranges": {"Revenue": "Sheet1!$B$5:$F$5"}
        }
    """
    # Load twice: once for formulas, once for computed values
    wb_formulas = load_workbook(filepath, data_only=False)
    wb_values = load_workbook(filepath, data_only=True)

    result = {
        "filename": Path(filepath).name,
        "sheets": {},
        "named_ranges": {}
    }

    # Extract named ranges
    for name in wb_formulas.defined_names:
        defn = wb_formulas.defined_names[name]
        result["named_ranges"][name] = defn.attr_text

    # Process each sheet
    for sheet_name in wb_formulas.sheetnames:
        ws_formulas = wb_formulas[sheet_name]
        ws_values = wb_values[sheet_name]

        sheet_data = {
            "dimensions": ws_formulas.dimensions,
            "cells": {}
        }

        for row_idx, row in enumerate(ws_formulas.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                cell_ref = f"{get_column_letter(col_idx)}{row_idx}"
                formula_val = cell.value
                computed_val = ws_values[cell_ref].value

                # Skip empty cells unless requested
                if formula_val is None and computed_val is None and not include_empty:
                    continue

                cell_data = {}

                # Check if it's a formula
                if formula_val is not None and str(formula_val).startswith('='):
                    cell_data["formula"] = str(formula_val)
                    cell_data["value"] = _serialize_value(computed_val)
                    cell_data["type"] = "formula"
                elif formula_val is not None:
                    cell_data["value"] = _serialize_value(formula_val)
                    cell_data["type"] = _get_type(formula_val)
                else:
                    cell_data["value"] = None
                    cell_data["type"] = "empty"

                # Add formatting if requested
                if include_formatting:
                    fmt = _get_formatting(cell)
                    if fmt:
                        cell_data["format"] = fmt

                sheet_data["cells"][cell_ref] = cell_data

        result["sheets"][sheet_name] = sheet_data

    wb_formulas.close()
    wb_values.close()

    return result


def _get_type(value) -> str:
    """Determine the type of a cell value."""
    if value is None:
        return "empty"
    elif isinstance(value, bool):
        return "boolean"
    elif isinstance(value, (int, float)):
        return "number"
    elif isinstance(value, str):
        return "string"
    elif isinstance(value, (datetime, date, time)):
        return "datetime"
    else:
        return "other"


def _serialize_value(value):
    """Convert value to JSON-serializable format."""
    if value is None:
        return None
    elif isinstance(value, (str, int, float, bool)):
        return value
    elif isinstance(value, datetime):
        return value.isoformat()
    elif isinstance(value, date):
        return value.isoformat()
    elif isinstance(value, time):
        return value.isoformat()
    else:
        # Handle any other types (like DataTableFormula) by converting to string
        return str(value)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    filepath = sys.argv[1]
    output_file = None

    # Parse --output arg
    if "--output" in sys.argv:
        idx = sys.argv.index("--output")
        if idx + 1 < len(sys.argv):
            output_file = sys.argv[idx + 1]

    data = xlsx_to_json(filepath)
    output = json.dumps(data, indent=2, default=str)

    if output_file:
        Path(output_file).write_text(output)
        print(f"Written to {output_file}")
    else:
        print(output)
